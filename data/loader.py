"""خواندن فایل‌های اکسل CRM (Notes / Tasks) به‌صورت مستقل از ساختار ستون‌ها.

این ماژول هیچ فرضی درباره‌ی نام دقیق ستون‌ها ندارد؛ فقط سطرها را به لیستی
از دیکشنری (header -> value) تبدیل می‌کند. تشخیص و Mapping ستون‌ها در
data/mapper.py انجام می‌شود.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl


class ExcelLoadError(Exception):
    """خطای قابل‌نمایش به کاربر (پیام فارسی)."""


@dataclass
class LoadedSheet:
    file_name: str
    sheet_name: str
    headers: list[str]
    rows: list[dict[str, Any]]  # هر ردیف: header -> value خام

    @property
    def row_count(self) -> int:
        return len(self.rows)


def _pick_best_sheet(wb) -> Any:
    """در فایل‌های Export شده از Dynamics معمولاً یک 'hiddenSheet' هم وجود دارد
    که فقط شامل متادیتای Mapping و لیست مقادیر Option-Set است، نه داده‌ی واقعی.
    شیتی را انتخاب می‌کنیم که بیشترین تعداد سطر و بیشترین تعداد ستون را دارد
    و Hidden نیست (در صورت وجود چند گزینه)."""
    candidates = []
    for ws in wb.worksheets:
        try:
            row_count = ws.max_row or 0
            col_count = ws.max_column or 0
        except Exception:
            row_count, col_count = 0, 0
        is_hidden = getattr(ws, "sheet_state", "visible") != "visible"
        score = row_count * max(col_count, 1)
        candidates.append((score, is_hidden, ws))

    if not candidates:
        raise ExcelLoadError("فایل اکسل هیچ شیتی ندارد.")

    visible = [c for c in candidates if not c[1]]
    pool = visible if visible else candidates
    pool.sort(key=lambda c: c[0], reverse=True)
    return pool[0][2]


def load_excel(path: str | Path, max_rows: int | None = None) -> LoadedSheet:
    """فایل اکسل را می‌خواند و بهترین شیت را برمی‌گرداند.

    Raises:
        ExcelLoadError: با پیام فارسی قابل‌نمایش به کاربر.
    """
    path = Path(path)
    if not path.exists():
        raise ExcelLoadError(f"فایل «{path.name}» یافت نشد.")
    if path.suffix.lower() not in (".xlsx", ".xlsm"):
        raise ExcelLoadError(
            f"فرمت فایل «{path.name}» پشتیبانی نمی‌شود. لطفاً فایل با پسوند .xlsx انتخاب کنید."
        )

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ExcelLoadError(
            f"فایل «{path.name}» باز نشد. ممکن است فایل خراب یا قفل‌شده باشد.\nجزئیات فنی: {exc}"
        ) from exc

    try:
        ws = _pick_best_sheet(wb)
        row_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(row_iter)
        except StopIteration:
            raise ExcelLoadError(f"فایل «{path.name}» خالی است.")

        headers = [str(h).strip() if h is not None else f"ستون بدون‌نام {i+1}" for i, h in enumerate(header_row)]
        rows: list[dict[str, Any]] = []
        for i, raw_row in enumerate(row_iter):
            if max_rows is not None and i >= max_rows:
                break
            if raw_row is None or all(v is None for v in raw_row):
                continue
            record = {headers[j]: raw_row[j] for j in range(min(len(headers), len(raw_row)))}
            rows.append(record)

        if not rows:
            raise ExcelLoadError(
                f"در فایل «{path.name}» هیچ رکورد دارای داده‌ای یافت نشد (فقط ردیف عنوان موجود است)."
            )

        return LoadedSheet(file_name=path.name, sheet_name=ws.title, headers=headers, rows=rows)
    finally:
        wb.close()
