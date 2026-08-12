from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _write_table(ws, headers: list[str], rows: list[list], rtl: bool = True):
    ws.sheet_view.rightToLeft = rtl
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r, row in enumerate(rows, start=2):
        for c, value in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.alignment = Alignment(horizontal="right" if rtl else "left", vertical="top", wrap_text=True)
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 24
    ws.freeze_panes = "A2"


def export_excel(path: str, data: dict) -> None:
    """data کلیدهای زیر را می‌پذیرد (هرکدام اختیاری؛ اگر وجود نداشته باشد آن شیت ساخته نمی‌شود):
    summary, cases, experts, criteria, comparison, data_quality, ai_analysis, suspicious
    هرکدام: {"headers": [...], "rows": [[...], ...]}"""
    wb = Workbook()
    first = True
    sheet_titles = {
        "summary": "Summary",
        "cases": "Case Analysis",
        "experts": "Expert Analysis",
        "criteria": "Criteria Scores",
        "comparison": "Period Comparison",
        "data_quality": "Data Quality",
        "ai_analysis": "AI Analysis",
        "suspicious": "Suspicious Cases",
    }
    for key, title in sheet_titles.items():
        section = data.get(key)
        if not section:
            continue
        ws = wb.active if first else wb.create_sheet()
        ws.title = title
        first = False
        _write_table(ws, section["headers"], section["rows"])
    if first:
        wb.active.title = "Summary"
    wb.save(path)
