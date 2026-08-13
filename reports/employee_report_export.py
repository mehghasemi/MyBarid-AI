"""Export گزارش عملکرد اختصاصی یک کارشناس به یک فایل اکسل تک‌شیتی و
خوانا، آماده ارسال مستقیم به همان کارشناس یا استفاده در جلسه Feedback."""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

TITLE_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
SECTION_FILL = PatternFill(start_color="EAF1F8", end_color="EAF1F8", fill_type="solid")
GOOD_FILL = PatternFill(start_color="E5F6EE", end_color="E5F6EE", fill_type="solid")
BAD_FILL = PatternFill(start_color="FBE9E9", end_color="FBE9E9", fill_type="solid")
WHITE_BOLD = Font(color="FFFFFF", bold=True, size=13)
BOLD = Font(bold=True)
RIGHT = Alignment(horizontal="right", vertical="top", wrap_text=True)


def export_expert_report_excel(path: str, report: dict) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "گزارش عملکرد"
    ws.sheet_view.rightToLeft = True
    for col, w in zip("ABCDE", [26, 20, 14, 14, 40]):
        ws.column_dimensions[col].width = w

    row = 1

    def write_title(text: str):
        nonlocal row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        c = ws.cell(row=row, column=1, value=text)
        c.fill = TITLE_FILL
        c.font = WHITE_BOLD
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 26
        row += 2

    def write_section(text: str):
        nonlocal row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        c = ws.cell(row=row, column=1, value=text)
        c.fill = SECTION_FILL
        c.font = BOLD
        c.alignment = RIGHT
        row += 1

    def write_row(*values, fill=None):
        nonlocal row
        for i, v in enumerate(values, start=1):
            c = ws.cell(row=row, column=i, value=v)
            c.alignment = RIGHT
            if fill:
                c.fill = fill
        row += 1

    def blank():
        nonlocal row
        row += 1

    if not report.get("has_data"):
        write_title(f"گزارش عملکرد: {report.get('expert', '')}")
        write_row(report.get("message", "داده‌ای برای این کارشناس موجود نیست."))
        wb.save(path)
        return

    s = report["summary"]
    unit_label = "Task" if report["unit"] == "task" else "Case"

    write_title(f"گزارش عملکرد اختصاصی — {report['expert']}")
    write_row(f"بازه ارزیابی: {report['period_label']}   |   واحد ارزیابی: {unit_label}")
    blank()

    write_section("خلاصه عملکرد")
    write_row(f"تعداد {unit_label} بررسی‌شده", s["case_count"])
    write_row("امتیاز Objective (Rule-Based)", s["avg_objective"])
    write_row("امتیاز AI", s["avg_ai"])
    write_row("امتیاز نهایی (Final Score)", s["avg_final"])
    write_row("تغییر نسبت به دوره قبل", s["change_vs_previous"])
    write_row("میانگین تیم در همین دوره", s["team_average"])
    write_row("فاصله با میانگین تیم", s["vs_team_average"])
    blank()

    ms = report["management_summary"]
    write_section("جمع‌بندی مدیریتی")
    write_row("وضعیت کلی", ms["overall_status"])
    write_row("مهم‌ترین نقطه قوت", ms["top_strength"])
    write_row("مهم‌ترین نقطه ضعف", ms["top_weakness"])
    write_row("اولویت اول بهبود", ms["top_priority"])
    write_row("پیشنهاد تمرکز دوره بعد", ms["next_focus"])
    blank()

    write_section("کارنامه معیارها (ضعیف‌ترین ابتدا)")
    write_row("معیار", "دسته", "میانگین امتیاز", "تعداد مورد بررسی‌شده", fill=SECTION_FILL)
    for item in report["scorecard"]:
        write_row(item["criterion"], item["category"], item["avg_score"], item["count"])
    blank()

    write_section("نقاط قوت")
    if report["strengths"]:
        write_row("معیار", "میانگین امتیاز", "تعداد", "نمونه‌ها (شماره — امتیاز)", fill=GOOD_FILL)
        for st in report["strengths"]:
            samples = "؛ ".join(f"{c['case_number'] or c['case_key']} ({c['score']})" for c in st["sample_cases"])
            write_row(st["criterion"], st["avg_score"], st["count"], samples)
    else:
        write_row("داده کافی برای شناسایی نقطه قوت مستند وجود ندارد.")
    blank()

    write_section("نقاط قابل بهبود (اولویت‌بندی‌شده)")
    if report["weaknesses"]:
        write_row("معیار", "اولویت", "میانگین امتیاز", "نمونه‌ها (شماره — امتیاز)", fill=BAD_FILL)
        for w in report["weaknesses"]:
            samples = "؛ ".join(f"{c['case_number'] or c['case_key']} ({c['score']})" for c in w["sample_cases"])
            write_row(w["criterion"], w["priority"], w["avg_score"], samples)
    else:
        write_row("ضعف قابل‌توجهی شناسایی نشد.")
    blank()

    write_section("Feedback (قابل ارائه مستقیم به کارشناس)")
    for line in report["feedback_lines"] or ["داده کافی برای تولید Feedback مستند وجود ندارد."]:
        write_row(line)
    blank()

    write_section("برنامه بهبود پیشنهادی")
    if report["action_plan"]:
        write_row("اولویت", "تمرکز", "دسته", "هدف", fill=SECTION_FILL)
        for a in report["action_plan"]:
            write_row(a["priority"], a["focus"], a["category"], a["target"])
    else:
        write_row("در حال حاضر اقدام اضطراری خاصی پیشنهاد نمی‌شود.")

    wb.save(path)
